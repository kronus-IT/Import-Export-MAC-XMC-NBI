#!/usr/bin/env python

##############################################################
# written by Markus Nikulski
#            mnikulski@extremenetworks.com
#            08 Mar 2022
# Edited by Martin Flammia
#            29 Mar 2022
#            mflammia@kronusit.co.uk
##############################################################

from queue import Empty
import time
import XMC_NBI
from openpyxl import Workbook, load_workbook
import os
import inquirer
import re

##############################################################
xmcServerIp     = '192.168.100.100'
xmcSecret       = 'ee7b07e0-d97c-4199-bfcc-434c0515fff9'
session         = None
baseMacAddress  = '00:00:11:11:22:20'
xmcClientID     = 'FIeAHVZuGm'
##############################################################
def login():
    global session
    session = XMC_NBI.XMC_NBI(xmcServerIp,xmcClientID,xmcSecret)
    if session.error:
        print( "ERROR: '%s'" % session.message )
        exit()

##############################################################
def getAllMACs():
    mac_list = session.getMacAddresses()
    if session.error:
        print( "ERROR: get MAC addresses failed '%s'" % session.message )
    else:
        return mac_list
        #print("INFO: found %s MACs" % len(mac_list))
        #for mac, group in sorted( mac_list.items() ):
            #print (( "\t%s => %s" % (mac,group) ) )

##############################################################
def getMAC(mac):
    data = session.getMacAddress(mac)
    if session.error:
        print( "ERROR: get MAC address failed '%s'" % session.message )
    else:
        if data == None:
            print( "INFO: MAC %s not exists" % mac )
        else:
            print("INFO: MAC %s exists in group %s with description '%s'" % (mac,data['groups'],data['groupDescription']) )
    
##############################################################
def addMAC(mac, group):
    session.addMacAddress(mac, group, 'added by test script')
    if session.error:
        print( "ERROR: add MAC address failed '%s'" % session.message )
    else:
        print("INFO: add MAC %s" % mac)
    
##############################################################
def delMAC(mac, group):
    session.delMacAddress(mac, group)
    if session.error:
        print( "ERROR: delete MAC address failed '%s'" % session.message )
    else:
        print("INFO: delete MAC %s" % mac)
    
##############################################################

def create_excel(all_macs,unique_groups):
    list_of_macs = list(all_macs.keys())
    total_number_of_macs = len(list_of_macs)
    total_number_of_groups = len(unique_groups)
    #Workbook is dynamically created by just calling the import workbook class
    wb = Workbook()
    #Workbook name
    dest_filename = 'MACs-In-Groups.xlsx'
    #Worbook is always created with ast least one worksheet. This will get the active worksheet
    ws = wb.active
    #Worksheets will be labelled by default sheet, sheet1, sheet2. This will change the title
    ws.title = "List-MACs-By-Group"
    #This creates a the top row of AP items starting one in from right
    unique_item = 1
    col_names = {}
    for unique_name in unique_groups:
        header_to_write = ws.cell(row=1, column=unique_item) #This writes the headers in row 1, where unique_item increases by 1
        col_names[unique_name] = unique_item #This creates a dictionary of column names column numbers
        header_to_write.value = unique_name
        unique_item += 1
    #Loop through the list of MAC addresses : Group disctionary values
    #Where the MAC addresses group matches the group title in first row, add the MAC address
    #Interate through each group at a time add MAC addresses
    unique_row = 2
    for unique_col_name, unique_group_col_number in col_names.items():
        for mac, group in all_macs.items(): #This loops through all the MACs getting MAC and Group value
            #if unique_group == group: #This matches on the unique_group where the group is the same
            if unique_col_name == group:
                cell_to_write = ws.cell(row=unique_row, column=unique_group_col_number)
                cell_to_write.value = mac
                unique_row += 1
        unique_row = 2
    #This saves the workbook
    wb.save(dest_filename)

##############################################################

def create_excel(all_macs,unique_groups):
    list_of_macs = list(all_macs.keys())
    total_number_of_macs = len(list_of_macs)
    total_number_of_groups = len(unique_groups)
    #Workbook is dynamically created by just calling the import workbook class
    wb = Workbook()
    #Workbook name
    dest_filename = 'MACs-In-Groups.xlsx'
    #Worbook is always created with ast least one worksheet. This will get the active worksheet
    ws = wb.active
    #Worksheets will be labelled by default sheet, sheet1, sheet2. This will change the title
    ws.title = "List-MACs-By-Group"
    #This creates a the top row of AP items starting one in from right
    unique_item = 1
    col_names = {}
    for unique_name in unique_groups:
        header_to_write = ws.cell(row=1, column=unique_item) #This writes the headers in row 1, where unique_item increases by 1
        col_names[unique_name] = unique_item #This creates a dictionary of column names column numbers
        header_to_write.value = unique_name
        unique_item += 1
    #Loop through the list of MAC addresses : Group disctionary values
    #Where the MAC addresses group matches the group title in first row, add the MAC address
    #Interate through each group at a time add MAC addresses
    unique_row = 2
    for unique_col_name, unique_group_col_number in col_names.items():
        for mac, group in all_macs.items(): #This loops through all the MACs getting MAC and Group value
            #if unique_group == group: #This matches on the unique_group where the group is the same
            if unique_col_name == group:
                cell_to_write = ws.cell(row=unique_row, column=unique_group_col_number)
                cell_to_write.value = mac
                unique_row += 1
        unique_row = 2
    #This saves the workbook
    wb.save(dest_filename)

##############################################################

def read_excel(xlsx_workbook, xlsx_sheets):
    #wb = load_workbook(filename = xlsx_workbook)
    ws = xlsx_workbook[xlsx_sheets[0]]
    unique_groups = {}
    all_macs = {}
    device_mac = []
    items_in_a_column = {}
    items_list = []
    row_count = {}
    unique_col_number = 1
    for cell in ws['1']: #This grabs everything that is in row or coloumn defined, in this case row 2, but could be column A
        unique_groups[cell.value] = unique_col_number #This creates a dictionary that has end-system group name as key and column number as value
        unique_col_number += 1
        #print(cell.value)
    #Create a dictionary that has the MAC address as the key and the group as the value
    number_of_unique_groups = len(unique_groups)
    all_macs = {}
    #iterate through each of the unique groups
    #for each group iterate through the matching row

    for cell in ws['1']: #This grabs everything that is in row defined, in this case row 1 which is where the headers are
        for col in ws.iter_cols(min_col=cell.column):#This iterates down the column as per the value cell.column and grabs all the cells in each row
            for row_cell in col: #This cycles through each cell in the row per column
                if row_cell.value is None: #This skips any cells that return a None value
                    continue
                else:
                    items_list.append(row_cell.value) #This creates a list of all the cell rows in the specific column
            items_in_a_column[cell.column] = items_list #This creates a dictionaty of the column number (key), and number of rows in that columns as a list against that key
            #This grabs the first entry in the list of entries in the coloum row, which is the header = items_list[0]
            #Then get the len of the list of rows in that same column = len(items_in_a_column[cell.column]
            #Wich then creates a dictionary of = column header name : number of rows
            #The minus 1 removes the count for the header in the total number of rows
            row_count.update({items_list[0]:len(items_in_a_column[cell.column]) - 1})
            items_list = []

    for unique_group, unique_group_col_id in unique_groups.items(): #This iterates through unique_groups getting unique group (key) and the unique group column ID (value)
        for cell in ws['1']: #This grabs everything that is in row defined, in this case row 1 which is where the headers are    
            if cell.value == unique_group: #This loops through the unique groups (cell.value) and matches when matching with a unique group name
                for cell_row in range(1,row_count[unique_group]): #This takes the total number of unique groups loops through the rows in each one.
                    d = ws.cell(row=cell_row + 1,column=unique_group_col_id) #This takes the row value against the coloumn ID
                    if d.value is not None: #This skips past any cell that has a value of None.
                        #print(d.value)
                        all_macs[d.value] = unique_group
    xlsx_workbook.close()
    return all_macs
    # Close the workbook after reading
    #wb.close()
    xlsx_workbook.close()

##############################################################

def write_mac (all_macs):
    for mac, group in all_macs.items(): #This loops through the dictionary of MACs + Group
        session.addMacAddress(mac,group) #This adds the MAC to the group on XMC
    return True

##############################################################

def get_unique_groups (all_macs):
    unique_groups = set() #This creates a set to store all the unqiue groups into
    for value in all_macs.values(): #This iterates through all the MACs creating a list of unique grousp
        unique_groups.add(value)
    return unique_groups

##############################################################

def create_groups(all_macs):
    unique_groups = set() #This creates a set to store all the unqiue groups into
    current_groups = []
    for value in all_macs.values(): #This iterates through all the MACs + Group dict all_macs, creating a list of unique groups
        unique_groups.add(value) #This adds each of the groups to set making sure they are unique
    current_groups.append(session.getESGroups()) #This reads all the currently existing groups in XMC
    list_of_xmc_groups = current_groups[0]
    list_of_unique_xls_groups = list(unique_groups)
    groups_that_do_not_exist_in_xmc = list(set(list_of_unique_xls_groups) - set(list_of_xmc_groups))
    for group in groups_that_do_not_exist_in_xmc: #This loops through the existing XMC groups
        unique_group_input = input("This group '{}' does not exist in XMC, select 'y' to create it, 'n' to continue, or 'x' to exit.".format(group))
        if unique_group_input == 'y':
            session.createGroup(group,'MAC') # The creation requires the type to be passed, MAC in this case
        elif unique_group_input == 'n':
            continue
        else:
            exit()

##############################################################

def select_file():
    files_xlsx = [] #This will hold a list of all files ending in extentsion .xlsx
    eos_vlan_config = {}
    current_path = os.getcwd() #This obtains the current full path like c:\users\martin\python
    file_ext = r".xlsx" #This determines the file format to search for
    for file in os.listdir(current_path): #This loops through the current path finding all files that end in .xlsx
        if file.endswith(file_ext):
            print(os.path.join("", file))
            files_xlsx.append(os.path.join("", file)) #And keeps appending them to the files_xlsx list
    #This sets up the question and the list of .xlsx files based on the files_xlsx list
    what_file = [inquirer.List('xlxs-files', message = "Which xlsx file below are you using as a template", choices = files_xlsx,)]
    #This initiates presenting the list of choices, and then stores the file name as a dictionary xlsx-files:<filename>
    xlsx_workbook_name = inquirer.prompt(what_file)
    #This loads the workbook into the xlsx_workbook based on the key 'xlsx-files' in xlsx_workbook_name
    xlsx_workbook = load_workbook(filename=xlsx_workbook_name['xlxs-files'])
    #This obtains a list of the worksheets as stored in the xlsx_workbook
    xlsx_sheets = xlsx_workbook.sheetnames
    #This sets up the question and the list of worksheets based on the xlxs-sheets list
    what_worksheet = [inquirer.List('xlxs-sheets', message = "Which sheet will you be using for interface and VLAN configuration", choices = xlsx_sheets,)]
    #This initiates presenting the list of choices, and then stores the worksheet name as a dictionary xlsx-sheets:<worksheet name>
    xlsx_worksheet_name = inquirer.prompt(what_worksheet)
    return xlsx_workbook, xlsx_sheets

##############################################################

def mac_check (all_macs):
    '''This checks the format of the MAC address, has to be 8 otects seperated by a - or a :'''
    for mac in all_macs.keys():
        if re.match("[0-9a-f]{2}([-:]?)[0-9a-f]{2}(\\1[0-9a-f]{2}){4}$", mac.upper()):
            continue
        else:
            print ("This MAC address {} format is incorrect. The MAC address should look similar to this xx:xx:xx:xx:xx:. Please correct and try again.".format(mac))
            exit()

##############################################################

def mac_duplicate (xlsx_macs):
    xmc_macs = getAllMACs()
    formated_xmc_macs = {}
    for xmc_mac_key, xmc_mac_value in xmc_macs.items():
        upper_xmc_mac = xmc_mac_key.upper()
        formated_xmc_macs.update({upper_xmc_mac:xmc_mac_value})
    remaining_macs = dict(set(xlsx_macs.items()) - set(formated_xmc_macs.items()) ) #This compares MACs from XLSX and from XMC and returns the remaining difference
    #value = { k : xlsx_macs[k] for k in set(xlsx_macs) - set(xmc_macs) }
    print ("These following MACs are going to be added as they do not currently exist in XMC")
    for key, value in remaining_macs.items():
        print (key, " => ", value)
    continute_install_macs = input("If you are happy with these, press 'y' to continue or 'x' to exit. ")
    if continute_install_macs == 'y':
        return remaining_macs
        print ("\n")
    else:
        exit()

login()

choice_export_import = input ("Would you like to export MACs [e], or Import MACs [i], choose [e] or [i]: ")
if choice_export_import == 'e':
    all_macs = getAllMACs() #This gets all the MAC's from XMC and creates a dictionary of MAC:Group
    unique_groups = get_unique_groups(all_macs) #This gets a list of unique groups by getting all the groups from all_macs and putting them into a set
    create_excel(all_macs,unique_groups) #This creates the spreedsheet of all the MACs, based its group which is added as a header based on the unique groups 
elif choice_export_import == 'i':
    xlsx_workbook,xlsx_sheets = select_file() #This uses inquirer to select xlsx and worksheet
    all_macs = read_excel(xlsx_workbook,xlsx_sheets) #This reads all the MAC's and creates a dictionary called all_macs that is formed of MAC:Group
    mac_check(all_macs) #This checks the MAC address has been formated correctly from the spreedsheet import
    remaining_macs = mac_duplicate(all_macs) #This returns only non duplicates by comparing xls MAC's with XMC MAC's, and prompts if you want to continue
    create_groups(remaining_macs) #This compares the groups in xls against those in XMC. If the group does not exist, you are prompted and creates in XMC
    write_mac(remaining_macs)
